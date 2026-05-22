"""utils_export.py
Excel 내보내기 및 데이터 출력 유틸리티

사용 예:
    import streamlit as st
    from utils_export import export_to_excel, render_download_button

    excel_bytes = export_to_excel(df, "문의목록")
    render_download_button(excel_bytes, "문의목록.xlsx")
"""
import io
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from helpers import get_logger

logger = get_logger(__name__)


def export_to_excel(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    *,
    title: str = "",
    include_summary: bool = False,
    number_columns: Optional[List[str]] = None,
) -> bytes:
    """DataFrame을 서식이 적용된 Excel 파일(bytes)로 변환한다.

    Args:
        df: 내보낼 데이터프레임
        sheet_name: 시트 이름
        title: 시트 상단에 표시할 제목 (빈 문자열이면 생략)
        include_summary: True이면 하단에 숫자 컬럼 합계 행 추가
        number_columns: 천 단위 콤마 서식을 적용할 컬럼 이름 목록

    Returns:
        Excel 파일의 바이트 데이터
    """
    if df.empty:
        df = pd.DataFrame({"정보": ["데이터가 없습니다"]})

    buf = io.BytesIO()
    start_row = 0

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if title:
            # 제목 행을 위한 빈 DataFrame
            header_df = pd.DataFrame()
            header_df.to_excel(writer, sheet_name=sheet_name, startrow=0)
            ws = writer.sheets[sheet_name]
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=2, column=1, value=f"출력일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            start_row = 3

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

        ws = writer.sheets[sheet_name]

        # 숫자 컬럼에 천 단위 서식 적용
        if number_columns:
            for col_name in number_columns:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1  # 1-based
                    for row_idx in range(start_row + 2, start_row + 2 + len(df)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.number_format = "#,##0"

        # 컬럼 너비 자동 조정
        from openpyxl.utils import get_column_letter

        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if not df.empty else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # 합계 행
        if include_summary and number_columns:
            summary_row = start_row + 2 + len(df)
            ws.cell(row=summary_row, column=1, value="합계")
            for col_name in number_columns:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    try:
                        total = pd.to_numeric(df[col_name], errors="coerce").sum()
                        cell = ws.cell(row=summary_row, column=col_idx, value=total)
                        cell.number_format = "#,##0"
                    except Exception:
                        pass

    buf.seek(0)
    logger.info(f"Excel export: {sheet_name} ({len(df)} rows)")
    return buf.getvalue()


def export_multiple_sheets(
    sheets: Dict[str, pd.DataFrame],
) -> bytes:
    """여러 시트를 하나의 Excel 파일로 내보낸다.

    Args:
        sheets: {시트이름: DataFrame, ...}

    Returns:
        Excel 파일 바이트 데이터
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df.empty:
                df = pd.DataFrame({"정보": ["데이터가 없습니다"]})
            safe_name = name[:31]  # Excel 시트 이름 최대 31자
            df.to_excel(writer, sheet_name=safe_name, index=False)
    buf.seek(0)
    logger.info(f"Multi-sheet export: {list(sheets.keys())}")
    return buf.getvalue()


def render_download_button(
    excel_bytes: bytes,
    filename: str = "export.xlsx",
    label: str = "📥 Excel 다운로드",
):
    """Streamlit 다운로드 버튼을 렌더링한다.

    Note: 이 함수는 streamlit 컨텍스트 안에서 호출해야 한다.
    """
    import streamlit as st

    st.download_button(
        label=label,
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
